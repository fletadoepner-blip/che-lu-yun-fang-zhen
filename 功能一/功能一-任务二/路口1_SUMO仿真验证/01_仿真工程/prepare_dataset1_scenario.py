"""Build a SUMO four-leg intersection directly from 路口数据/1 workbook."""

from __future__ import annotations

import os
import subprocess
from pathlib import Path
from xml.etree.ElementTree import Element, ElementTree, SubElement

import openpyxl


PROJECT = Path(__file__).resolve().parent
SCENARIO = PROJECT / "scenario"
WORKBOOK = SCENARIO / "demo_1_流量和配时方案.xlsx"
CALIBRATED_ADAPTIVE_GREEN = (42, 30, 47, 21)


def write_xml(root: Element, path: Path) -> None:
    ElementTree(root).write(path, encoding="utf-8", xml_declaration=True)


def build_network(netconvert: str) -> None:
    nodes = Element("nodes")
    for node_id, x, y, node_type in (
        ("W", -220, 0, "priority"), ("E", 220, 0, "priority"),
        ("N", 0, 220, "priority"), ("S", 0, -220, "priority"),
        ("J1", 0, 0, "traffic_light"),
    ):
        attributes = {"id": node_id, "x": str(x), "y": str(y), "type": node_type}
        if node_id == "J1":
            attributes["radius"] = "15"
        SubElement(nodes, "node", **attributes)
    edges = Element("edges")
    for edge_id, source, target in (
        ("W_in", "W", "J1"), ("E_in", "E", "J1"), ("N_in", "N", "J1"), ("S_in", "S", "J1"),
        ("W_out", "J1", "W"), ("E_out", "J1", "E"), ("N_out", "J1", "N"), ("S_out", "J1", "S"),
    ):
        SubElement(edges, "edge", id=edge_id, **{"from": source, "to": target, "numLanes": "3", "speed": "13.89"})
    connections = Element("connections")
    # Indexes define the state-string positions in the Excel-derived signal program.
    # SUMO lane 0 is the rightmost lane of travel. Therefore, for every
    # approach: left turn uses lane 2, through uses lane 1, right turn uses lane 0.
    movement_data = (
        ("W_in", 2, "N_out", 2), ("W_in", 1, "E_out", 1), ("W_in", 0, "S_out", 0),
        ("E_in", 2, "S_out", 2), ("E_in", 1, "W_out", 1), ("E_in", 0, "N_out", 0),
        ("N_in", 0, "W_out", 0), ("N_in", 1, "S_out", 1), ("N_in", 2, "E_out", 2),
        ("S_in", 0, "E_out", 0), ("S_in", 1, "N_out", 1), ("S_in", 2, "W_out", 2),
    )
    for index, (source, source_lane, target, target_lane) in enumerate(movement_data):
        SubElement(connections, "connection", **{
            "from": source, "to": target, "fromLane": str(source_lane), "toLane": str(target_lane),
            "tl": "J1", "linkIndex": str(index),
        })
    write_xml(nodes, SCENARIO / "dataset_1.nod.xml")
    write_xml(edges, SCENARIO / "dataset_1.edg.xml")
    write_xml(connections, SCENARIO / "dataset_1.con.xml")
    subprocess.run([
        netconvert, "--node-files", str(SCENARIO / "dataset_1.nod.xml"),
        "--edge-files", str(SCENARIO / "dataset_1.edg.xml"),
        "--connection-files", str(SCENARIO / "dataset_1.con.xml"),
        "--output-file", str(SCENARIO / "dataset_1.net.xml"), "--no-turnarounds", "true",
        "--no-warnings", "true",
    ], check=True)


def controlled_connections() -> dict[tuple[str, str], dict[str, int]]:
    """Read netconvert's real link ordering instead of assuming source XML order."""
    root = ElementTree(file=SCENARIO / "dataset_1.net.xml").getroot()
    result: dict[tuple[str, str], dict[str, int]] = {}
    for connection in root.findall("connection"):
        if connection.get("tl") != "J1":
            continue
        result[(connection.get("from"), connection.get("to"))] = {
            "index": int(connection.get("linkIndex")),
            "lane": int(connection.get("fromLane")),
        }
    if len(result) != 12:
        raise RuntimeError(f"Expected 12 controlled movements, found {len(result)}")
    return result


def signal_state(connections, movements: dict[tuple[str, str], str], yellow: bool = False) -> str:
    """Build an RYG state from actual link indexes returned by netconvert."""
    states = ["r"] * len(connections)
    for movement, color in movements.items():
        index = connections[movement]["index"]
        states[index] = "y" if yellow else color
    return "".join(states)


def validate_phase_axis(connections, state: str, expected_sources: set[str]) -> None:
    """Fail generation if a phase accidentally opens perpendicular approaches."""
    active_sources = {
        source for (source, _), data in connections.items()
        if state[data["index"]] in "Gg"
    }
    if not active_sources.issubset(expected_sources):
        raise RuntimeError(
            f"Unsafe phase: expected {sorted(expected_sources)}, got {sorted(active_sources)}"
        )


def build_demand_and_signal() -> None:
    workbook = openpyxl.load_workbook(WORKBOOK, data_only=True)
    flow_sheet = workbook["流量数据"]
    timing_sheet = workbook["信号配时数据"]
    # Rows 4--11 are the eight 15-minute early-peak observations; columns follow E, W, N, S and L, T, R.
    approaches = ("E", "W", "N", "S")
    outputs = {
        "E": ("S_out", "W_out", "N_out"), "W": ("N_out", "E_out", "S_out"),
        "N": ("W_out", "S_out", "E_out"), "S": ("E_out", "N_out", "W_out"),
    }
    connections = controlled_connections()
    routes = Element("routes")
    movement_totals = {(approach, turn): 0 for approach in approaches
                       for turn in ("left", "through", "right")}
    SubElement(routes, "vType", id="passenger", accel="2.6", decel="4.5", sigma="0.5",
               tau="1.2", minGap="2.5", length="5", maxSpeed="13.89")
    for approach in approaches:
        for turn, target in zip(("left", "through", "right"), outputs[approach]):
            SubElement(routes, "route", id=f"{approach}_{turn}", edges=f"{approach}_in {target}")
    for interval, row in enumerate(range(4, 12)):
        begin, end = interval * 900, (interval + 1) * 900
        for approach_index, approach in enumerate(approaches):
            for turn_index, turn in enumerate(("left", "through", "right")):
                volume = int(flow_sheet.cell(row, 3 + approach_index * 3 + turn_index).value)
                movement_totals[(approach, turn)] += volume
                target = outputs[approach][turn_index]
                lane = connections[(f"{approach}_in", target)]["lane"]
                SubElement(routes, "flow", id=f"peak_{interval}_{approach}_{turn}", type="passenger",
                           route=f"{approach}_{turn}", begin=str(begin), end=str(end), number=str(volume),
                           departLane=str(lane), departSpeed="random")
    write_xml(routes, SCENARIO / "dataset_1.rou.xml")

    green = [int(timing_sheet.cell(row, 5).value) for row in range(3, 7)]
    yellow = [int(timing_sheet.cell(row, 6).value) for row in range(3, 7)]
    all_red = [int(timing_sheet.cell(row, 7).value) for row in range(3, 7)]
    # Values are keyed by physical movement, not fragile state-string offsets.
    # Straight/right and left turns use separate phases. Opposing left turns use
    # yielding green so SUMO resolves their geometric conflict safely.
    phase_movements = (
        {("E_in", "W_out"): "G", ("E_in", "N_out"): "g",
         ("W_in", "E_out"): "G", ("W_in", "S_out"): "g"},
        {("E_in", "S_out"): "g", ("W_in", "N_out"): "g"},
        {("N_in", "S_out"): "G", ("N_in", "W_out"): "g",
         ("S_in", "N_out"): "G", ("S_in", "E_out"): "g"},
        {("N_in", "E_out"): "g", ("S_in", "W_out"): "g"},
    )
    phase_demands = (
        sum(movement_totals[(approach, turn)] for approach in ("E", "W")
            for turn in ("through", "right")),
        sum(movement_totals[(approach, "left")] for approach in ("E", "W")),
        sum(movement_totals[(approach, turn)] for approach in ("N", "S")
            for turn in ("through", "right")),
        sum(movement_totals[(approach, "left")] for approach in ("N", "S")),
    )
    green_budget = sum(green)
    adaptive_green = list(CALIBRATED_ADAPTIVE_GREEN)
    override = os.environ.get("ADAPTIVE_GREEN_TIMES")
    if override:
        adaptive_green = [int(value) for value in override.split(",")]
    if len(adaptive_green) != 4 or sum(adaptive_green) != green_budget:
        raise ValueError(f"Adaptive greens must contain four values summing to {green_budget}")
    additional = Element("additional")
    for program_id, green_times in (("dataset_fixed", green), ("dataset_adaptive", adaptive_green)):
        logic = SubElement(additional, "tlLogic", id="J1", type="static", programID=program_id, offset="0")
        for index in range(4):
            movements = phase_movements[index]
            green_state = signal_state(connections, movements)
            expected_sources = {"E_in", "W_in"} if index < 2 else {"N_in", "S_in"}
            validate_phase_axis(connections, green_state, expected_sources)
            SubElement(logic, "phase", duration=str(green_times[index]), state=green_state)
            SubElement(logic, "phase", duration=str(yellow[index]),
                       state=signal_state(connections, movements, yellow=True))
            SubElement(logic, "phase", duration=str(all_red[index]), state="r" * 12)
    write_xml(additional, SCENARIO / "dataset_1.add.xml")

    configuration = Element("configuration")
    input_node = SubElement(configuration, "input")
    SubElement(input_node, "net-file", value="dataset_1.net.xml")
    SubElement(input_node, "route-files", value="dataset_1.rou.xml")
    SubElement(input_node, "additional-files", value="dataset_1.add.xml")
    write_xml(configuration, SCENARIO / "dataset_1.sumocfg")


def prepare(netconvert: str) -> Path:
    if not WORKBOOK.exists():
        raise FileNotFoundError(f"Missing 路口数据/1 workbook: {WORKBOOK}")
    SCENARIO.mkdir(exist_ok=True)
    build_network(netconvert)
    build_demand_and_signal()
    return SCENARIO / "dataset_1.sumocfg"
